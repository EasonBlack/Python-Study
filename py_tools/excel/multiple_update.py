
import openpyxl

theFile = openpyxl.load_workbook('/Users/eason/Downloads/xx.xlsx')
 
productMap = {}
hasDiffList = []

print(theFile.sheetnames)
currentSheet = theFile['零售商品导入模板']
for row in range(5, currentSheet.max_row):
  name = "{}{}".format("M", row)
  category =  "{}{}".format("B", row)
  code = "{}{}".format("D", row)
  desname = "{}{}".format("N", row)
  strName = currentSheet[name].value
  strCode = currentSheet[code].value
  strCategory = currentSheet[category].value
  if productMap.get(strName) and  productMap.get(strName) != strCategory :
    thingIndex =  hasDiffList.index(strName) if strName in hasDiffList else -1
    if thingIndex == -1 :
      hasDiffList.append(strName)
  elif  productMap.get(strName, "") == "" : 
    productMap[strName] = strCategory
  # if strCode != '' :
  #   strName = strName.split("-")[0]
  #   currentSheet[desname] = strName
 
print(hasDiffList)
# theFile.save('/Users/eason/Downloads/xx1.xlsx')
