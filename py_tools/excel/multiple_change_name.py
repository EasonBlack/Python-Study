
import openpyxl

theFile = openpyxl.load_workbook('/Users/eason/Downloads/xx.xlsx')
 
productMap = {}
hasDiffList = []

print(theFile.sheetnames)
currentSheet = theFile['xxx']
for row in range(1, currentSheet.max_row):
  name = "{}{}".format("A", row)
  desname = "{}{}".format("N", row)
  strName = currentSheet[name].value

  if strName != ""  and strName is not None:
    print(strName)
    strNameList = strName.split(" ")
    if len(strNameList) > 1:
      strNameList.pop()
    print(strNameList)
    currentSheet[desname] = " ".join(strNameList)
    print(" ".join(strNameList))
 
# print(hasDiffList)
theFile.save('/Users/eason/Downloads/xx1.xlsx')
